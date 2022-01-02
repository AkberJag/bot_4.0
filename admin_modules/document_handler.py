import os
import random
import traceback
import openpyxl

from time import sleep
from django.db.models import Q
from telebot.types import ReplyKeyboardRemove
from admin_modules.weekly_pad import process_weeklyXL_pad, send_pad_weekly
from sms.send_sms import daily_sms
from datetime import datetime, timedelta
from bot_modules.config import msg_auth_key
from admin_modules import process_bill_text
from bot_modules.config import ADMIN, MILMA_NAME
from bot_modules.jason_handler import add_to_json
from admin_modules.weekly_mky import process_weeklyXL_mky, send_mky_weekly
from bot_modules.send_msg import send_warnings_to_admin, send_messages
from bot_modules.warning_msgs import (
    daily_bill_upload,
    daily_bill,
    daily_bill_success,
    weekly_pad_upload,
    daily_sms_success,
)
from bot_modules.bot_keyboards import date_keyboard


def make_XL_from_db():
    from db.models import User

    workbook = openpyxl.Workbook()
    wb_sheet = workbook.worksheets[0]
    for user in User.objects.all().values_list():
        wb_sheet.append(
            [
                str(user[3]),  # milma id
                user[1],  # name
                str(user[4]),  # Phone
                str(user[5]),  # pw
                str(user[2]),  # tg id
            ]
        )
    workbook.save("contacts.xlsx")


# remove unwanted rows and cloums from the contact xl
def sanitize_XL(workbook, file_path):
    contacts_excel_row_length = 5
    wb_sheet = workbook.worksheets[0]
    # delete empty rows
    for _ in range(int(wb_sheet.max_row)):

        # delete colums wich contains unwanted data
        for i in range(contacts_excel_row_length + 1, wb_sheet.max_column):
            wb_sheet.delete_cols(i)

        # check if the password and telegram Id column empty or not
        # if empty, set them a value
        for i, row in enumerate(wb_sheet.iter_rows()):
            row_items = list(x.value for x in row)
            # check password cell
            if row_items[3] == None:
                wb_sheet[i + 1][3].value = str(
                    random.randint(1000, 9999)
                )  # actual value get save in DB
                row_items[3] = "1234"  # dummy value

            # check telegram id cell
            if row_items[4] == None:
                wb_sheet[i + 1][4].value = "0"
                row_items[4] = "0"

        # if pw and telgram id is there but still empty cells
        # delete them
        for i, row in enumerate(wb_sheet.iter_rows()):
            # check none for first 5 colums only
            if None in list(x.value for x in row)[:4]:
                wb_sheet.delete_rows(i + 1, 1)

        # save the new excel file
        workbook.save(file_path)


# here we save all XL and text files from telegram to disk
def save_file_from_tg(message, bot, file_name):
    fileID = message.document.file_id
    file_info = bot.get_file(fileID)
    downloaded_file = bot.download_file(file_info.file_path)
    with open(file_name, "wb") as new_file:
        new_file.write(downloaded_file)


# move all unused text and XL files to disk
def move_old_files(file_name, file_type):
    try:
        os.renames(
            file_name,  # old file name
            f"{os.getcwd()}/"  # current directory
            + "other_files/"  # common folder for all old files
            + f"{file_type}/"  # file type - contact file or daily bill etc
            + f"{datetime.now().strftime('%Y')}/"  # year
            + f"{datetime.now().strftime('%m')}/"  # month
            + f"{(datetime.now() + timedelta(hours=5, minutes=30)).strftime('%d-%m-%Y %I.%M.%S.%f %p')}_"[
                :100  # get only 100 charactes of time
            ]  # current time
            + file_name,
        )
    except:
        pass


def send_tg_msg_and_save_to_json_csv(bot, file_data):
    from db.models import User, DailyBill

    # get each users data then save to db and send via tg
    delivery_success = 0
    sms_delivery_success = 0

    daily_msg_details = []

    failed_users = ""
    sms_failed_users = ""

    daily_bill_pk = 0
    delivery_details = ""
    for data in file_data["datas"][1:]:
        print(data[file_data["datas"][0].index("Pro Id")])
        try:
            # save daily bill details to db
            daily_bill_pk = DailyBill.objects.update_or_create(
                milma_id=data[file_data["datas"][0].index("Pro Id")],
                Qty=data[file_data["datas"][0].index("Qty")],
                Fat=data[file_data["datas"][0].index("FAT")],
                Clr=data[file_data["datas"][0].index("CLR")],
                Snf=data[file_data["datas"][0].index("SNF")],
                Rate=data[file_data["datas"][0].index("Rate")],
                Amount=data[file_data["datas"][0].index("Amount")],
                AM_Pm=file_data["time"][0],
                Date=datetime.strptime(
                    datetime.strptime(
                        file_data["date"]
                        .lower()
                        .replace("am", "")
                        .replace("pm", "")
                        .strip(),
                        "%d-%m-%Y",
                    ).strftime("%d-%b-%Y"),
                    "%d-%b-%Y",
                ),
            )
        except Exception as e:
            print(e)

        # select a registered user with milma id and a telegram id != 0
        user = User.objects.filter(
            ~Q(telegram_id=0),
            ~Q(telegram_id=None),
            milma_id=int(data[file_data["datas"][0].index("Pro Id")]),
        )

        # if user is exist send telegram bill
        if user:
            sleep(0.05)
            delivery_details = send_messages(
                bot,
                daily_bill.format(
                    date=f"{file_data['date']}",
                    name=user.values()[0]["name"].title(),
                    id=user.values()[0]["milma_id"],
                    qty=data[file_data["datas"][0].index("Qty")],
                    clr=data[file_data["datas"][0].index("CLR")],
                    fat=data[file_data["datas"][0].index("FAT")],
                    snf=data[file_data["datas"][0].index("SNF")],
                    rs=data[file_data["datas"][0].index("Rate")],
                    total=data[file_data["datas"][0].index("Amount")],
                ),
                user.values()[0]["telegram_id"],
            )

            # if the message send succesfully in telegram
            # save the id to delete in future
            if delivery_details["reply"] is not None:
                delivery_success += 1
                daily_msg_details.append(
                    [
                        delivery_details["reply"].chat.id,
                        delivery_details["reply"].message_id,
                        daily_bill_pk[0].pk,
                    ]  # save the chat id and msg id for deleting
                )

            # if the message is failed in telegram
            # notify the admins
            else:
                failed_users += f"{user.values()[0]['name'].title()} ({user.values()[0]['milma_id']})\n"
                print(
                    f"{user.values()[0]['name'].title()} ({user.values()[0]['milma_id']}) - {delivery_details.get('error')}"
                )

        # send SMS if the msg failed in telegram
        else:
            pass

    # add the message and all recived users msg id and chat id
    # to delete a messsage in future.
    add_to_json(
        "{}_{}".format(
            file_data["file_type"],
            f"{file_data['date']} {file_data['time']}",  # date with AM or PM
        ),
        daily_msg_details,
    )

    # send success message to admin
    send_warnings_to_admin(
        bot,
        daily_bill_success.format(delivery_success)
        + f'\n\n{daily_sms_success.format(sms_delivery_success) if sms_delivery_success > 0 else ""}',
    )

    # check if any message is failed, then send their details
    if failed_users != "":
        nl = "\n\n"
        send_warnings_to_admin(
            bot,
            f'{"<b>Telegram message failed for these users:</b>{}{}".format(nl,failed_users[:3900]) if failed_users != "" else "" }'
            + f'\n\n{"Sending SMS failed for these users:{}".format(sms_failed_users) if sms_failed_users != "" else ""}',
        )


def send_tg_msg_and_save_to_json_txt(bot, file_data):
    from db.models import User, DailyBill

    # get each users data then save to db and send via tg
    delivery_success = 0
    sms_delivery_success = 0

    daily_msg_details = []

    failed_users = ""
    sms_failed_users = ""

    daily_bill_pk = 0
    delivery_details = ""
    for data in file_data["datas"]:
        try:
            # save daily bill details to db
            daily_bill_pk = DailyBill.objects.update_or_create(
                milma_id=data[0],
                Qty=data[2],
                Fat=data[3],
                Clr=data[4],
                Snf=data[5],
                Rate=data[6],
                Amount=data[7],
                AM_Pm=file_data["time"][0],
                Date=datetime.strptime(file_data["date"], "%d-%b-%Y"),
            )
        except:
            pass

        # select a registered user with milma id and a telegram id != 0
        user = User.objects.filter(
            ~Q(telegram_id=0), ~Q(telegram_id=None), milma_id=int(data[0])
        )

        # if user is exist send telegram bill
        if user:
            sleep(0.05)
            delivery_details = send_messages(
                bot,
                daily_bill.format(
                    date=f"{file_data['date']} {file_data['time']}",
                    name=user.values()[0]["name"].title(),
                    id=user.values()[0]["milma_id"],
                    qty=data[2],
                    clr=data[4],
                    fat=data[3],
                    snf=data[5],
                    rs=data[6],
                    total=data[7],
                ),
                user.values()[0]["telegram_id"],
            )

            # if the message send succesfully in telegram
            # save the id to delete in future
            if delivery_details["reply"] is not None:
                delivery_success += 1
                daily_msg_details.append(
                    [
                        delivery_details["reply"].chat.id,
                        delivery_details["reply"].message_id,
                        daily_bill_pk[0].pk,
                    ]  # save the chat id and msg id for deleting
                )

            # if the message is failed in telegram
            # notify the admins
            else:
                failed_users += f"{user.values()[0]['name'].title()} ({user.values()[0]['milma_id']})\n"
                print(
                    f"{user.values()[0]['name'].title()} ({user.values()[0]['milma_id']}) - {delivery_details.get('error')}"
                )

        # send SMS if the msg failed in telegram
        else:
            user = User.objects.filter(milma_id=int(data[0]))

            # send the sms only if there is a AUTH_KEY and the user exist on the db
            if msg_auth_key and user:
                sms_delivery_details = daily_sms(
                    data,
                    name=f"{user.values()[0]['name'][:15].title()} ({user.values()[0]['milma_id']})",
                    phone=user.values()[0]["phone"],
                    date=f"{file_data['date']} {file_data['time']}",
                )

                if sms_delivery_details["reply"] is not None:
                    sms_delivery_success += 1
                else:
                    sms_failed_users += f"{user.values()[0]['name'].title()} ({user.values()[0]['milma_id']})\n"
                    print(
                        f"{user.values()[0]['name'].title()} ({user.values()[0]['milma_id']}) - {delivery_details.get('error')}"
                    )

    # add the message and all recived users msg id and chat id
    # to delete a messsage in future.
    add_to_json(
        "{}_{}".format(
            file_data["file_type"],
            f"{file_data['date']} {file_data['time']}",  # date with AM or PM
        ),
        daily_msg_details,
    )

    # send success message to admin
    send_warnings_to_admin(
        bot,
        daily_bill_success.format(delivery_success)
        + f'\n\n{daily_sms_success.format(sms_delivery_success) if sms_delivery_success > 0 else ""}',
    )

    # check if any message is failed, then send their details
    if failed_users != "":
        nl = "\n\n"
        send_warnings_to_admin(
            bot,
            f'{"<b>Telegram message failed for these users:</b>{}{}".format(nl,failed_users[:3900]) if failed_users != "" else "" }'
            + f'\n\n{"Sending SMS failed for these users:{}".format(sms_failed_users) if sms_failed_users != "" else ""}',
        )


# get the documents from telegram
# only txt and xlsx files are processed
def document_from_telegram(message, bot):
    from db.models import User

    # daily bill text
    if ".txt" in message.document.file_name:

        # save text file to temp location
        # fmt: off
        file_path = str(bot.get_file(message.document.file_id).file_path).replace("documents/", "")
        text_file_id = message.message_id
        # fmt: on

        # save the daily text file to disk
        save_file_from_tg(message, bot, file_path)

        # all users data from input text file.
        text_file_data = process_bill_text.get_txt_file_data(file_path)
        if text_file_data == 0:
            # send bill detail msg to admins
            send_warnings_to_admin(bot, "This is not a proper daily bill file")

            # forward this file to me
            bot.forward_message(ADMIN, message.chat.id, text_file_id)

            # move unused file to a directory
            move_old_files(file_path, "old_daily_bill")

        else:
            # extra info for delete json
            text_file_data["file_type"] = "daily_bill"

            # send bill detail msg to admins
            send_warnings_to_admin(
                bot,
                daily_bill_upload.format(
                    f"{text_file_data['date']} {text_file_data['time']}",
                    len(text_file_data["datas"]),
                ),
            )

            # move unused file to a directory
            move_old_files(file_path, "old_daily_bill")

            # send msg/SMS to users and save the data to delete json
            send_tg_msg_and_save_to_json_txt(bot, text_file_data)

    # daily bill from NIC (CSV file)
    elif ".csv" in message.document.file_name:
        # save text file to temp location
        # fmt: off
        file_path = str(bot.get_file(message.document.file_id).file_path).replace("documents/", "")
        text_file_id = message.message_id
        # fmt: on

        # save the daily text file to disk
        save_file_from_tg(message, bot, file_path)

        # all users data from csv file.
        csv_file_data = process_bill_text.get_csv_file_data(file_path)

        if csv_file_data == 0:
            # send bill detail msg to admins
            send_warnings_to_admin(bot, "This is not a proper daily bill file")

            # forward this file to me
            bot.forward_message(ADMIN, message.chat.id, text_file_id)

            # move unused file to a directory
            move_old_files(file_path, "old_daily_bill_csv")

        else:
            # extra info for delete json
            csv_file_data["file_type"] = "daily_bill"
            # move unused file to a directory
            move_old_files(file_path, "old_daily_bill_csv")

            # add date keyboard
            date_keyboard(bot, message)
            bot.register_next_step_handler(
                message, get_date_from_milma, bot, csv_file_data
            )

    # if the file is a contacts.xlsx file
    elif message.document.file_name == "contacts.xlsx":
        print(message.document.file_name)
        # save contacts.xlsx
        # file to temp location
        # fmt: off
        # get "telegram file name" not the exact file name
        file_path = str(bot.get_file(message.document.file_id).file_path).replace("documents/", "")
        # fmt: on

        # save the contacts XL file to disk
        save_file_from_tg(message, bot, file_path)

        # load and get datas from the excel file
        workbook = openpyxl.load_workbook(file_path)

        # clear unwanted datas from the excel file
        sanitize_XL(workbook, file_path)

        # load and get datas from the excel file
        workbook = openpyxl.load_workbook(file_path)
        wb_sheet = workbook.worksheets[0]
        for row in wb_sheet.iter_rows():
            try:
                # check if the user is already in the DB
                if not User.objects.filter(
                    Q(milma_id=row[0].value) | Q(phone=row[2].value)
                ):
                    User.objects.update_or_create(
                        milma_id=row[0].value,
                        name=str(row[1].value).capitalize(),
                        phone=row[2].value,
                        pword=1234 if row[3].value is None else row[3].value,
                        telegram_id=0 if row[4].value is None else row[4].value,
                    )
            except:
                pass

        # move unused file to a directory
        move_old_files(file_path, "old_contactsXL")

        send_warnings_to_admin(
            bot,
            f"✅ <b>EXCEL file updated successfully</b> ✅\n\nRegistered users: {User.objects.filter(~Q(telegram_id=0)).count()}\nNon registered users: {User.objects.filter(telegram_id=0).count()}",
        )

    # if the file is any other xl file
    elif ".xlsx" in message.document.file_name:
        if MILMA_NAME == "APCOS മുരിക്കാശ്ശേരി":
            print(f"{MILMA_NAME} another xl")
            # save contacts.xlsx
            # file to temp location
            # fmt: off
            # get "telegram file name" not the exact file name
            file_path = str(bot.get_file(message.document.file_id).file_path).replace("documents/", "")
            print(file_path)
            # fmt: on
            # this is to forward the file to admin is something goes wrong
            file_message_id = message.message_id

            # save the XL file to disk
            save_file_from_tg(message, bot, file_path)

            weekly_data = process_weeklyXL_mky(file_path)

            # send bill details sending msg to admins
            send_warnings_to_admin(
                bot,
                weekly_pad_upload.format(weekly_data["date"], len(weekly_data["data"])),
            )

            # move unused file to a directory
            move_old_files(file_path, "old_weekly_mky_XL")

            # if result is succeful send SMS to users
            if weekly_data["result"] == "Success":
                send_mky_weekly(bot, weekly_data)

            # if failed, warn admin
            else:
                # forward this file to me
                if message.chat.id != ADMIN:
                    bot.forward_message(ADMIN, message.chat.id, file_message_id)

                send_warnings_to_admin(
                    bot,
                    "<b>എന്തോ കുഴപ്പം ഉണ്ട്..</b>\n\nഎക്സൽ ഫയലിലെ വിവരങ്ങൾ പ്രോസസ്സ് ചെയ്യാൻ പറ്റിയില്ല!😢😢",
                )
                import pprint

                bot.send_message(ADMIN, pprint.pformat(weekly_data)[:4000])

        elif MILMA_NAME == "APCOS പടമുഖം":
            print(f"{MILMA_NAME} another xl")
            # save contacts.xlsx
            # file to temp location
            # fmt: off
            # get "telegram file name" not the exact file name
            file_path = str(bot.get_file(message.document.file_id).file_path).replace("documents/", "")
            print(file_path)
            # fmt: on
            # this is to forward the file to admin is something goes wrong
            file_message_id = message.message_id

            # save the XL file to disk
            save_file_from_tg(message, bot, file_path)

            weekly_data = process_weeklyXL_pad(file_path)

            # send bill details sending msg to admins
            send_warnings_to_admin(
                bot,
                weekly_pad_upload.format(weekly_data["date"], len(weekly_data["data"])),
            )

            # move unused file to a directory
            move_old_files(file_path, "old_weekly_pad_XL")

            # if result is succeful send message to users
            if weekly_data["result"] == "Success":
                send_pad_weekly(bot, weekly_data)

            # if failed, warn admin
            else:
                # forward this file to me
                if message.chat.id != ADMIN:
                    bot.forward_message(ADMIN, message.chat.id, file_message_id)

                send_warnings_to_admin(
                    bot,
                    "<b>എന്തോ കുഴപ്പം ഉണ്ട്..</b>\n\nഎക്സൽ ഫയലിലെ വിവരങ്ങൾ പ്രോസസ്സ് ചെയ്യാൻ പറ്റിയില്ല!😢😢",
                )
                import pprint

                bot.send_message(ADMIN, pprint.pformat(weekly_data)[:4000])


def get_date_from_milma(message, bot, csv_file_data):
    if "Cancel" in message.text:
        bot.send_message(message.chat.id, "🔥", reply_markup=ReplyKeyboardRemove())
    else:
        try:
            datetime.strptime(message.text, "%d-%m-%Y %p")
            csv_file_data["date"] = message.text
            csv_file_data["time"] = "AM" if "am" in str(message.text).lower() else "PM"

            # send bill detail msg to admins
            send_warnings_to_admin(
                bot,
                daily_bill_upload.format(
                    csv_file_data["date"],
                    len(
                        csv_file_data["datas"][1:]
                    ),  # [1:] is here because first line is heading
                ),
            )

            # # send msg/SMS to users and save the data to delete json
            send_tg_msg_and_save_to_json_csv(bot, csv_file_data)

        except ValueError:
            print(traceback.format_exc())

            send_warnings_to_admin(
                bot, f"{message.text} ഈ തീയതി തെറ്റാണ് വീണ്ടും ശ്രമിക്കു."
            )
            date_keyboard(bot, message)
            bot.register_next_step_handler(
                message, get_date_from_milma, bot, csv_file_data
            )
