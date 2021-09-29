import openpyxl
import traceback

from time import sleep
from bot_modules.jason_handler import add_to_json
from bot_modules.warning_msgs import weekly_pad_success, weekly_msg_pad
from bot_modules.send_msg import send_warnings_to_admin, send_to_usrs_img


def process_weeklyXL_mky(file_path):
    mky_weekly_heading_start = 0
    try:
        print("\nmky weekly\n")
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        wb_sheet = workbook.worksheets[0]

        mky_weekly_heading_start = -1
        week_range = "തിയതി നൽകിയിട്ടില്ല"
        weekly_bill_details = []
        for i, row in enumerate(wb_sheet.iter_rows(), start=1):
            # if mno and name present, that is our heading
            if "header" and "name" in [str(cell.value).lower() for cell in row]:
                # heading
                mky_weekly_heading_start = i

        if mky_weekly_heading_start > 5:
            return {
                "result": None,
                "data": "0",
                "date": "Error",
                "error": f"Greater than 5 {mky_weekly_heading_start}\n{traceback.format_exc()[:3997]}",
            }

        elif mky_weekly_heading_start == -1:
            return {
                "result": None,
                "data": "0",
                "date": "Error",
                "error": f"No heading find error weekly {mky_weekly_heading_start}\n{traceback.format_exc()[:3997]}",
            }
        week_range = "തിയതി നൽകിയിട്ടില്ല"
        try:
            if "20" in str(
                [
                    str(cell.value).lower()
                    for cell in wb_sheet[mky_weekly_heading_start - 1]
                ]
            ):
                heading_list = [
                    str(cell.value).lower()
                    for cell in wb_sheet[mky_weekly_heading_start - 1]
                    if cell
                ][0]
                week_range = heading_list[heading_list.find(":") + 1 :]
        except:
            pass

        # heading of the excel file
        XL_HEADING = [
            str(cell.value).lower()
            for cell in wb_sheet[mky_weekly_heading_start]
            if cell.value
        ]

        # get data of each user from the excel file
        for row in wb_sheet.iter_rows(mky_weekly_heading_start + 1):
            row_values = [cell.value for cell in row][: len(XL_HEADING)]
            if None not in row_values:
                weekly_bill_details.append(
                    {
                        "bank_acc": row_values[0],
                        "member_name": str(row_values[1]).capitalize(),
                        "milma_member_id": row_values[2],
                        "weekly_amount": row_values[3],
                    }
                )
        return {
            "result": "Success",
            "data": weekly_bill_details,
            "heading": XL_HEADING,
            "file_type": "Mky_Weekly",
            "date": week_range,
        }
    except:
        return {
            "result": "Error",
            "data": "0",
            "date": "Error",
            "error": f"Some error {mky_weekly_heading_start}\n{traceback.format_exc()[:3997]}",
        }


def send_mky_weekly(bot, weekly_data):
    from db.models import User
    from django.db.models import Q

    delivery_success = 0
    weekly_msg_details = []

    for data in weekly_data["data"]:

        # select a registered user with milma id and a telegram id != 0
        user = User.objects.filter(
            ~Q(telegram_id="0"),
            ~Q(telegram_id=None),
            milma_id=int(data["milma_member_id"]),
        ).values()

        # if user is exist send telegram bill
        if user:
            # sleep(0.05)
            print(user)
            if str(user[0]["milma_id"]) == str(data["milma_member_id"]):
                # send weekly message to users
                # introduce a small delay
                delivery_details = send_to_usrs_img(
                    bot,
                    f"{weekly_msg_pad}\n{str(weekly_data['date']).replace('to', 'മുതൽ')} വരെ\n\n{data['member_name']}, {data['milma_member_id']}\nഅക്കൗണ്ട് നമ്പർ: XXXX{data['bank_acc'][-4:]}<b>\n\n₹{data['weekly_amount']} രൂപ ബാങ്ക് അക്കൗണ്ടിൽ നിക്ഷേപിച്ചിരിക്കുന്നു </b>",
                    user[0]["telegram_id"],
                )

                if delivery_details is not None:
                    delivery_success += 1
                    weekly_msg_details.append(
                        [
                            delivery_details.chat.id,
                            delivery_details.message_id,
                        ]  # save the chat id and msg id for deleting
                    )

    # send success message to admin
    send_warnings_to_admin(bot, weekly_pad_success.format(delivery_success))
    delivery_success = 0

    # add the message and all recived users msg id and chat id
    # to delete a messsage in future.
    add_to_json(
        f"{weekly_data['file_type']} Week: {weekly_data['date']}",
        weekly_msg_details,
    )
