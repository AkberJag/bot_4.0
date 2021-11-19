import openpyxl
import traceback

from time import sleep
from bot_modules.config import msg_auth_key
from bot_modules.jason_handler import add_to_json
from bot_modules.warning_msgs import (
    weekly_pad_success,
    weekly_msg_pad,
    daily_sms_success,
)
from bot_modules.send_msg import send_warnings_to_admin, send_to_usrs_img


import openpyxl, traceback

from sms.send_sms import weekly_sms


def process_weeklyXL_pad(file_path):
    pad_weekly_heading_start = 0
    try:
        print("\npad weekly\n")
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        wb_sheet = workbook.worksheets[0]

        pad_weekly_heading_start = -1
        week_range = "തിയതി നൽകിയിട്ടില്ല"
        weekly_bill_details = []
        for i, row in enumerate(wb_sheet.iter_rows(), start=1):
            # if mno and name present, that is our heading
            if "amount" and "name" in [str(cell.value).lower() for cell in row]:
                # heading
                pad_weekly_heading_start = i
                break

        if pad_weekly_heading_start > 5:
            return {
                "result": None,
                "data": "0",
                "date": "Error",
                "error": f"Greater than 5 - {pad_weekly_heading_start}\n{traceback.format_exc()[:3997]}",
            }

        elif pad_weekly_heading_start == -1:
            return {
                "result": None,
                "data": "0",
                "date": "Error",
                "error": f"No heading find error weekly - {pad_weekly_heading_start}\n{traceback.format_exc()[:3997]}",
            }
        try:
            heading = [
                str(cell.value).lower()
                for cell in wb_sheet[pad_weekly_heading_start - 1]
                if cell.value
            ]
            if "20" in str(heading):
                week_range = (
                    heading[0][heading[0].find(":") + 1 :].strip().replace("to", "മുതൽ")
                    + " വരെ"
                )
        except:
            pass

        # heading of the excel file
        XL_HEADING = [
            str(cell.value).lower()
            for cell in wb_sheet[pad_weekly_heading_start]
            if str(cell.value).replace(".", "") and cell.value
        ]

        # get data of each user from the excel file
        for row in wb_sheet.iter_rows(pad_weekly_heading_start + 1):
            row_values = [0 if cell.value == None else cell.value for cell in row][
                : len(XL_HEADING)
            ]
            if None not in row_values and row_values[0] != 0:
                weekly_bill_details.append(row_values)
        return {
            "result": "Success",
            "data": weekly_bill_details,
            "heading": XL_HEADING,
            "file_type": "pad_Weekly",
            "date": week_range,
        }
    except:
        return {
            "result": "Error",
            "data": "0",
            "date": "Error",
            "error": f"Some error {pad_weekly_heading_start}\n{traceback.format_exc()[:3997]}",
        }


def send_pad_weekly(bot, weekly_data):
    from db.models import User
    from django.db.models import Q

    delivery_success = 0
    weekly_msg_details = []

    sms_delivery_success = 0

    sms_failed_users = ""

    for data in weekly_data.get("data"):
        weekly_msg = ""

        # select a registered user with milma id and a telegram id != 0
        user = User.objects.filter(
            ~Q(telegram_id="0"),
            ~Q(telegram_id=None),
            milma_id=int(data[0]),
        ).values()

        # if user is exist send telegram bill
        if user:
            # introduce a small delay
            sleep(0.05)
            if str(user[0]["milma_id"]) == str(int(data[0])):
                # send weekly message to users
                weekly_msg += weekly_data.get("date") + "\n\n"
                weekly_msg += (
                    f"<b>{user[0]['name']} {data[0]}</b>\n\n"  # Name and milma no
                )
                for i, heading in enumerate(weekly_data.get("heading")[2:], start=2):
                    weekly_msg += f"{heading.capitalize()}: "
                    weekly_msg += (
                        f"{data[i]:.2f}\n" if type(data[i]) == float else f"{data[i]}\n"
                    )

                delivery_details = send_to_usrs_img(
                    bot,
                    f"{weekly_msg_pad}\n{weekly_msg}",
                    user[0]["telegram_id"],
                )

                if delivery_details is not None:
                    delivery_success += 1
                    weekly_msg_details.append(
                        [
                            delivery_details.chat.id,
                            delivery_details.message_id,
                        ]  # save the chat id and msg id for deleting
                    )
        # send SMS if the msg failed in telegram
        else:
            user = User.objects.filter(milma_id=int(data[0]))
            # send the sms only if there is a AUTH_KEY and the user exist on the db
            if msg_auth_key and user:
                sms_delivery_details = weekly_sms(
                    data,
                    user.values(),
                    weekly_data.get("heading"),
                    weekly_data.get("date"),
                )

                if sms_delivery_details["reply"] == "success":
                    sms_delivery_success += 1
                else:
                    sms_failed_users += f"{user.values()[0]['name'].title()} ({user.values()[0]['milma_id']})\n"

    # send success message to admin
    send_warnings_to_admin(
        bot,
        weekly_pad_success.format(delivery_success)
        + f'\n\n{daily_sms_success.format(sms_delivery_success) if sms_delivery_success > 0 else ""}',
    )
    delivery_success = 0

    # add the message and all recived users msg id and chat id
    # to delete a messsage in future.
    add_to_json(
        f"{weekly_data['file_type']} Week: {weekly_data['date']}",
        weekly_msg_details,
    )
